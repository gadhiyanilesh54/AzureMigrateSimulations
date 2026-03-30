"""RVTools CSV/XLSX import — convert RVTools exports into discovery data.

Parses RVTools vInfo, vCPU, vMemory, vDisk, vNetwork, vHost sheets and converts
them into the project's internal discovery format.
"""

from __future__ import annotations

import csv
import io
import re
from typing import Any


# RVTools field → internal field mapping
_VINFO_MAP = {
    "VM": "name",
    "Name": "name",
    "Powerstate": "power_state",
    "Power state": "power_state",
    "CPUs": "num_cpus",
    "Num CPUs": "num_cpus",
    "Memory": "memory_mb",
    "Memory MB": "memory_mb",
    "NICs": "nic_count",
    "Provisioned MB": "provisioned_mb",
    "Provisioned MiB": "provisioned_mb",
    "In Use MB": "in_use_mb",
    "In Use MiB": "in_use_mb",
    "OS according to the configuration file": "guest_os",
    "OS according to the VMware Tools": "guest_os_tools",
    "Guest OS": "guest_os",
    "Datacenter": "datacenter",
    "Cluster": "cluster",
    "Host": "host",
    "Folder": "folder",
    "Resource pool": "resource_pool",
    "DNS Name": "guest_hostname",
    "Annotation": "annotation",
    "HW version": "hardware_version",
    "Hardware Version": "hardware_version",
    "VM UUID": "instance_uuid",
    "UUID": "instance_uuid",
}


def import_rvtools_csv(
    csv_content: str,
    source_filename: str = "rvtools_export.csv",
) -> dict[str, Any]:
    """Import RVTools CSV data and convert to discovery format.

    Accepts the content of a single CSV file (typically vInfo export).
    Returns a discovery data dict compatible with vcenter_discovery.json.
    """
    reader = csv.DictReader(io.StringIO(csv_content))

    vms: list[dict[str, Any]] = []
    hosts_seen: dict[str, dict] = {}

    for row in reader:
        vm = _parse_vm_row(row)
        if vm and vm.get("name"):
            vms.append(vm)
            # Track hosts
            host_name = vm.get("host", "")
            if host_name and host_name not in hosts_seen:
                hosts_seen[host_name] = {
                    "name": host_name,
                    "datacenter": vm.get("datacenter", ""),
                    "cluster": vm.get("cluster", ""),
                }

    hosts = list(hosts_seen.values())

    # Extract unique networks, datastores, clusters, datacenters
    networks = sorted(set(v.get("network", "") for v in vms if v.get("network")))
    clusters = sorted(set(v.get("cluster", "") for v in vms if v.get("cluster")))
    datacenters = sorted(set(v.get("datacenter", "") for v in vms if v.get("datacenter")))

    return {
        "vms": vms,
        "hosts": hosts,
        "datastores": [],
        "networks": [{"name": n, "type": "VM Network"} for n in networks],
        "clusters": [{"name": c} for c in clusters],
        "datacenters": [{"name": d} for d in datacenters],
        "import_source": "rvtools",
        "import_filename": source_filename,
        "vm_count": len(vms),
        "host_count": len(hosts),
    }


def import_rvtools_xlsx(file_bytes: bytes, filename: str = "rvtools.xlsx") -> dict[str, Any]:
    """Import RVTools XLSX file.

    Requires openpyxl (optional dependency). Falls back to error if not installed.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "XLSX import requires openpyxl: pip install openpyxl"}

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Try to find vInfo sheet
    vinfo_sheet = None
    for name in wb.sheetnames:
        if name.lower().startswith("vinfo"):
            vinfo_sheet = wb[name]
            break

    if not vinfo_sheet:
        # Fall back to first sheet
        vinfo_sheet = wb[wb.sheetnames[0]]

    # Convert sheet to CSV-like dicts
    rows = list(vinfo_sheet.iter_rows(values_only=True))
    if not rows:
        return {"error": "Empty sheet", "vms": [], "hosts": []}

    headers = [str(h or "").strip() for h in rows[0]]
    vms: list[dict[str, Any]] = []
    hosts_seen: dict[str, dict] = {}

    for row in rows[1:]:
        row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        vm = _parse_vm_row(row_dict)
        if vm and vm.get("name"):
            vms.append(vm)
            host_name = vm.get("host", "")
            if host_name and host_name not in hosts_seen:
                hosts_seen[host_name] = {
                    "name": host_name,
                    "datacenter": vm.get("datacenter", ""),
                    "cluster": vm.get("cluster", ""),
                }

    wb.close()

    return {
        "vms": vms,
        "hosts": list(hosts_seen.values()),
        "datastores": [],
        "networks": [],
        "clusters": [],
        "datacenters": [],
        "import_source": "rvtools_xlsx",
        "import_filename": filename,
        "vm_count": len(vms),
        "host_count": len(hosts_seen),
    }


def _parse_vm_row(row: dict[str, Any]) -> dict[str, Any] | None:
    """Parse a single RVTools row into a VM dict."""
    vm: dict[str, Any] = {}

    for rvtools_key, internal_key in _VINFO_MAP.items():
        if rvtools_key in row and row[rvtools_key] is not None:
            vm[internal_key] = row[rvtools_key]

    if not vm.get("name"):
        return None

    # Normalize numeric fields
    vm["num_cpus"] = _to_int(vm.get("num_cpus", 0))
    vm["memory_mb"] = _to_int(vm.get("memory_mb", 0))

    # Calculate disk size
    provisioned = _to_float(vm.get("provisioned_mb", 0))
    vm["total_disk_gb"] = round(provisioned / 1024, 1) if provisioned else 0
    vm.pop("provisioned_mb", None)
    vm.pop("in_use_mb", None)

    # Normalize power state
    ps = str(vm.get("power_state", "")).lower()
    if "on" in ps:
        vm["power_state"] = "poweredOn"
    elif "off" in ps:
        vm["power_state"] = "poweredOff"
    elif "suspend" in ps:
        vm["power_state"] = "suspended"

    # Classify guest OS family
    guest_os = str(vm.get("guest_os", vm.get("guest_os_tools", ""))).lower()
    if "windows" in guest_os:
        vm["guest_os_family"] = "windows"
    elif any(x in guest_os for x in ["linux", "ubuntu", "centos", "rhel", "debian", "suse", "oracle"]):
        vm["guest_os_family"] = "linux"
    else:
        vm["guest_os_family"] = "other"

    # Ensure required fields
    vm.setdefault("vcenter_id", "rvtools-import")
    vm.setdefault("disks", [])
    vm.setdefault("nics", [])
    vm.setdefault("instance_uuid", "")

    return vm


def _to_int(val: Any) -> int:
    try:
        return int(float(str(val).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def _to_float(val: Any) -> float:
    try:
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0
